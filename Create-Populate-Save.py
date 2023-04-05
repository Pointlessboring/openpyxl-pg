########################
# Dabbling with Openpyxl
#
#
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = Workbook()
ws = wb.active
ws.title = "New Title"

# Creating some empty sheets
ws1 = wb.create_sheet("New#1")
ws2 = wb.create_sheet("New#2")
ws3 = wb.create_sheet("New#3")

# Printing sheetnames as a list
print(wb.sheetnames)

# Printing sheetnames
for sheet in wb:
    print(sheet.title)

# Filling our worksheet with cells and data.
NUMROWS = 10
NUMCOLS = 10

# Populate worksheet cells
for x in range(1,NUMROWS+1):
    for y in range(1,NUMCOLS+1):
        ws.cell(row=x, column=y, value = x*y)

# Inserting the header row
ws.insert_rows(1)

for y in range(1,NUMCOLS+1):
    ws.cell(row=1, column=y, value = f"COL{y}")

# Formatting the header row
for y in range(1,NUMCOLS+1):
    mycell = ws.cell(row=1, column=y)
    mycell.font = Font(bold=True)
    mycell.alignment = Alignment(horizontal="center", vertical="center")
    mycell.fill = PatternFill("solid", fgColor="DDDDDD")

# Set view to 200%
ws.sheet_view.zoomScale = 200

# Saving our workbook to disk
wb.save('MyFirstExcelFile.xlsx')
